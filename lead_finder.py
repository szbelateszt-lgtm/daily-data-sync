"""
Foodora Lead Finder - Új éttermek és élelmiszerboltok automatikus felderítése
Budapest + agglomeráció területén.

Adatforrások:
1. Google Places API (New) - hivatalos, megbízható
2. crt.sh - új .hu domainek vendéglátós kulcsszavakkal
3. Google News RSS - friss hírek új helyekről
4. We Love Budapest gasztro - frissen nyílt helyek cikkei
5. Funzine.hu gasztro - frissen nyílt helyek cikkei
6. Time Out Budapest - frissen nyílt helyek cikkei

A script naponta egyszer fut le GitHub Actions-ben.
Az új találatokat e-mailben küldi el.
"""

import os
import re
import io
import json
import smtplib
import ssl
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from pathlib import Path
import requests
import feedparser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# KONFIGURÁCIÓ
# ============================================================

GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")
EMAIL_FROM = os.environ.get("EMAIL_FROM", "")
EMAIL_TO = os.environ.get("EMAIL_TO", "")
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD", "")  # Gmail app password
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465

# Keresési területek: Budapest 23 kerülete + agglomeráció főbb városai
# Minden terület egy kör középponttal és sugárral (méterben)
# ============================================================
# KERESÉSI KONFIGURÁCIÓ
# ============================================================
# Háromféle terület-típus:
#  1) Belvárosi sűrű rács: kis (350m) körök, DISTANCE rangsor → az új,
#     kevés véleményű helyek is beférnek a 20-as limitbe, mert a kör
#     elég kicsi és a legközelebbieket adja vissza (nem a legnépszerűbbeket)
#  2) Külső Budapest: nagy körök, RELEVANCE + konyha-specifikus keresések
#  3) Agglomeráció: nagy körök, csak alap keresések

# Query-listák
BASIC_QUERIES = ["étterem", "kávézó", "élelmiszerbolt", "pékség", "bisztró", "kifőzde"]
GRID_QUERIES = ["étterem", "kávézó", "bisztró", "kifőzde", "pékség"]
OUTER_CUISINE = [
    "vietnami étterem", "thai étterem", "ramen", "sushi",
    "kebab", "burger", "pizzéria", "kínai étterem",
]

# Belvárosi rács középpontjai (~600m-enként, 350m sugarú körök)
# Lefedi az V., VI., VII., VIII., IX. kerület magját, ahol a legtöbb új hely nyílik.
_BELVAROS_GRID = [
    (47.5033, 19.0441), (47.5033, 19.0521), (47.5033, 19.0601),
    (47.4979, 19.0441), (47.4979, 19.0521), (47.4979, 19.0601),
    (47.4925, 19.0441), (47.4925, 19.0521), (47.4925, 19.0601),
]

SEARCH_AREAS = []

# 1) Belvárosi sűrű rács
for _i, (_lat, _lng) in enumerate(_BELVAROS_GRID, 1):
    SEARCH_AREAS.append({
        "name": f"Belváros #{_i}",
        "qloc": "Budapest",        # ezt fűzzük a query-hez (a kört a locationBias adja)
        "lat": _lat, "lng": _lng, "radius": 350,
        "queries": GRID_QUERIES,
        "rank": "DISTANCE",
    })

# 2) Külső Budapest (nagy körök, relevancia + konyha-specifikus)
SEARCH_AREAS += [
    {"name": "Budapest észak", "qloc": "Budapest észak", "lat": 47.5500, "lng": 19.0700,
     "radius": 5000, "queries": BASIC_QUERIES + OUTER_CUISINE, "rank": "RELEVANCE"},
    {"name": "Budapest dél", "qloc": "Budapest dél", "lat": 47.4500, "lng": 19.0700,
     "radius": 5000, "queries": BASIC_QUERIES + OUTER_CUISINE, "rank": "RELEVANCE"},
    {"name": "Budapest kelet", "qloc": "Budapest kelet", "lat": 47.4900, "lng": 19.1500,
     "radius": 5000, "queries": BASIC_QUERIES + OUTER_CUISINE, "rank": "RELEVANCE"},
    {"name": "Budapest nyugat", "qloc": "Budapest nyugat", "lat": 47.5100, "lng": 19.0100,
     "radius": 3000, "queries": BASIC_QUERIES + OUTER_CUISINE, "rank": "RELEVANCE"},
]

# 3) Agglomeráció (nagy körök, csak alap keresések)
SEARCH_AREAS += [
    {"name": "Budaörs", "qloc": "Budaörs", "lat": 47.4626, "lng": 18.9580,
     "radius": 3000, "queries": BASIC_QUERIES, "rank": "RELEVANCE"},
    {"name": "Érd", "qloc": "Érd", "lat": 47.3950, "lng": 18.9050,
     "radius": 4000, "queries": BASIC_QUERIES, "rank": "RELEVANCE"},
    {"name": "Szentendre", "qloc": "Szentendre", "lat": 47.6669, "lng": 19.0760,
     "radius": 3000, "queries": BASIC_QUERIES, "rank": "RELEVANCE"},
    {"name": "Gödöllő", "qloc": "Gödöllő", "lat": 47.6000, "lng": 19.3550,
     "radius": 3000, "queries": BASIC_QUERIES, "rank": "RELEVANCE"},
    {"name": "Vecsés", "qloc": "Vecsés", "lat": 47.4100, "lng": 19.2700,
     "radius": 3000, "queries": BASIC_QUERIES, "rank": "RELEVANCE"},
    {"name": "Dunakeszi", "qloc": "Dunakeszi", "lat": 47.6350, "lng": 19.1400,
     "radius": 3000, "queries": BASIC_QUERIES, "rank": "RELEVANCE"},
]

# Hely típusok, amik érdekelnek minket (Foodora-szempontból releváns)
RELEVANT_TYPES = [
    "restaurant",
    "cafe",
    "bakery",
    "bar",
    "meal_takeaway",
    "meal_delivery",
    "convenience_store",
    "grocery_store",
    "supermarket",
    "food_store",
]

# Adattárolás
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
KNOWN_PLACES_FILE = DATA_DIR / "known_places.json"
KNOWN_DOMAINS_FILE = DATA_DIR / "known_domains.json"
KNOWN_WLB_FILE = DATA_DIR / "known_wlb_articles.json"
KNOWN_FUNZINE_FILE = DATA_DIR / "known_funzine_articles.json"
KNOWN_TIMEOUT_FILE = DATA_DIR / "known_timeout_articles.json"
KNOWN_GOVCENTER_FILE = DATA_DIR / "known_govcenter_uzletek.json"

# Govcenter.hu kerület-azonosítók (onk_id).
# A govcenter.hu publikus üzletnyilvántartást szolgáltat ezekhez a kerületekhez.
# Bővíthető, ahogy újabb kerületek onk_id-ját azonosítjuk.
# tip=1: bejelentéshez kötött kereskedelmi tevékenység
# tip=2: működési engedéllyel rendelkező üzletek
GOVCENTER_KERULETEK = [
    # Megerősített kerületek (irányítószám-alapon ellenőrizve)
    {"name": "Budapest I. (Várnegyed)", "onk_id": 544},
    {"name": "Budapest II. (Rózsadomb)", "onk_id": 463},
    {"name": "Budapest V. (Belváros)", "onk_id": 432},
    {"name": "Budapest VI. (Terézváros)", "onk_id": 182},
    {"name": "Budapest VII. (Erzsébetváros)", "onk_id": 647},
    {"name": "Budapest VIII. (Józsefváros)", "onk_id": 58},
    {"name": "Budapest IX. (Ferencváros)", "onk_id": 507},
    {"name": "Budapest XI. (Újbuda)", "onk_id": 465},
    {"name": "Budapest XII. (Hegyvidék)", "onk_id": 126},
    {"name": "Budapest XIV. (Zugló)", "onk_id": 563},
    {"name": "Budapest XV. (Rákospalota)", "onk_id": 169},
    {"name": "Budapest XVI. (Sashalom)", "onk_id": 10},
    {"name": "Budapest XVIII. (Pestszentlőrinc)", "onk_id": 462},
    {"name": "Budapest XIX. (Kispest)", "onk_id": 548},
    {"name": "Budapest XX. (Pesterzsébet)", "onk_id": 578},
    {"name": "Budapest XXI. (Csepel)", "onk_id": 564},
    {"name": "Budapest XXII. (Budafok-Tétény)", "onk_id": 583},
    {"name": "Budapest XXIII. (Soroksár)", "onk_id": 519},
    {"name": "Budapest XXIII/b (Soroksár)", "onk_id": 587},
    # Bizonytalan — az első futás megmutatja melyik kerület valójában
    {"name": "Budapest III.? (Óbuda?)", "onk_id": 1},
    {"name": "Budapest IV.? (Újpest?)", "onk_id": 7},
]

# ============================================================
# 1. GOOGLE PLACES API - ÚJ ÉS HAMAROSAN NYÍLÓ HELYEK
# ============================================================

def fetch_google_places():
    """
    Google Places API New - Text Search hívás minden területre.
    A FUTURE_OPENING státuszú helyek mellett a frissen megjelenteket is gyűjti.
    """
    if not GOOGLE_API_KEY:
        print("⚠️  GOOGLE_API_KEY nincs beállítva, kihagyom.")
        return []

    all_places = []
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": GOOGLE_API_KEY,
        "X-Goog-FieldMask": (
            "places.id,places.displayName,places.formattedAddress,"
            "places.businessStatus,places.openingDate,places.primaryType,"
            "places.googleMapsUri,places.websiteUri,"
            "places.nationalPhoneNumber,places.location"
        ),
    }

    for area in SEARCH_AREAS:
        queries = area["queries"]
        rank = area.get("rank", "RELEVANCE")
        for query in queries:
            body = {
                "textQuery": f"{query} {area['qloc']}",
                "locationBias": {
                    "circle": {
                        "center": {
                            "latitude": area["lat"],
                            "longitude": area["lng"],
                        },
                        "radius": area["radius"],
                    }
                },
                "maxResultCount": 20,
                "rankPreference": rank,
                "includeFutureOpeningBusinesses": True,
            }

            try:
                resp = requests.post(
                    "https://places.googleapis.com/v1/places:searchText",
                    headers=headers,
                    json=body,
                    timeout=30,
                )
                if resp.status_code != 200:
                    print(f"  ✗ Hiba {area['name']} / {query}: HTTP {resp.status_code} – {resp.text[:200]}")
                    continue
                data = resp.json()
                places = data.get("places", [])
                for p in places:
                    p["_search_area"] = area["name"]
                    p["_search_query"] = query
                all_places.extend(places)
                print(f"  ✓ {area['name']} / {query}: {len(places)} hely")
            except Exception as e:
                print(f"  ✗ Hiba {area['name']} / {query}: {e}")

    # Deduplikálás place_id alapján
    unique = {}
    for p in all_places:
        pid = p.get("id")
        if pid and pid not in unique:
            unique[pid] = p
    print(f"📍 Összesen {len(unique)} egyedi hely")
    return list(unique.values())


def filter_new_places(places, known_ids):
    """Csak azok a helyek érdekesek, amik még nem voltak ismertek."""
    new = []
    for p in places:
        pid = p.get("id")
        if pid and pid not in known_ids:
            new.append(p)
    return new


# ============================================================
# 2. crt.sh - ÚJ .hu DOMAIN REGISZTRÁCIÓK
# ============================================================

def fetch_new_domains():
    """
    Új SSL tanúsítványokat kérdez le a Certificate Transparency logból
    vendéglátós kulcsszavakkal. Minden új weboldal automatikusan kap SSL-t,
    így ez gyakorlatilag az új weboldalak listája.
    """
    keywords = [
        "etterem", "restaurant", "bisztro", "bistro", "kavezo",
        "pizza", "burger", "pekseg", "bakery", "kifozde",
        "kocsma", "bar", "elelmiszer", "grocery", "deli",
        "konyha", "kitchen", "food", "cafe", "trattoria",
    ]

    cutoff = (datetime.utcnow() - timedelta(days=2)).strftime("%Y-%m-%d")
    all_domains = set()

    headers = {
        "User-Agent": "Mozilla/5.0 (Foodora Lead Finder; +https://github.com/)"
    }
    for kw in keywords:
        url = f"https://crt.sh/?q=%25{kw}%25.hu&output=json"
        try:
            resp = requests.get(url, headers=headers, timeout=30)
            if resp.status_code != 200:
                print(f"  ⚠️  crt.sh '{kw}': HTTP {resp.status_code}")
                continue
            data = resp.json()
            for cert in data:
                # Csak az utóbbi 2 napban kiadott tanúsítványok
                if cert.get("entry_timestamp", "")[:10] < cutoff:
                    continue
                names = cert.get("name_value", "").split("\n")
                for name in names:
                    name = name.strip().lower()
                    # Csak .hu domainek, nem wildcard, nem aldomain
                    if (name.endswith(".hu")
                            and "*" not in name
                            and name.count(".") == 1):
                        all_domains.add(name)
            print(f"  ✓ '{kw}': eddig {len(all_domains)} unique domain")
        except Exception as e:
            print(f"  ✗ Hiba '{kw}': {e}")

    return list(all_domains)


# ============================================================
# 3. GOOGLE NEWS RSS - FRISS HÍREK
# ============================================================

def fetch_news():
    """Google News RSS keresés vendéglátós kulcsszavakkal."""
    queries = [
        "új étterem Budapest",
        "megnyílt Budapest étterem",
        "nyitás Budapest étterem",
        "új kávézó Budapest",
        "új bisztró Budapest",
        "új élelmiszerbolt Budapest",
    ]

    all_news = []
    # 7 napos cutoff (volt 2 napos, de túl szigorú volt és minden hírt eldobott)
    cutoff = datetime.utcnow() - timedelta(days=7)

    for q in queries:
        url = (
            f"https://news.google.com/rss/search?q={requests.utils.quote(q)}"
            f"&hl=hu&gl=HU&ceid=HU:hu"
        )
        try:
            feed = feedparser.parse(url)
            kept = 0
            for entry in feed.entries[:10]:
                # Próbáljuk megkapni a dátumot többféleképpen
                pub_dt = None
                pub_parsed = entry.get("published_parsed")
                if pub_parsed:
                    try:
                        pub_dt = datetime(*pub_parsed[:6])
                    except (TypeError, ValueError):
                        pub_dt = None

                # Ha nem tudjuk a dátumot, BEENGEDJÜK a hírt (jobb beengedni, mint elveszíteni)
                # Ha tudjuk és túl régi, kihagyjuk
                if pub_dt and pub_dt < cutoff:
                    continue

                all_news.append({
                    "title": entry.get("title", ""),
                    "link": entry.get("link", ""),
                    "published": entry.get("published", ""),
                    "source": entry.get("source", {}).get("title", "") if entry.get("source") else "",
                    "query": q,
                })
                kept += 1
            print(f"  ✓ '{q}': {len(feed.entries[:10])} talált, {kept} elfogadva")
        except Exception as e:
            print(f"  ✗ Hiba '{q}': {e}")

    # Deduplikálás link alapján
    seen = set()
    unique = []
    for n in all_news:
        if n["link"] not in seen:
            seen.add(n["link"])
            unique.append(n)
    return unique


# ============================================================
# GASZTROPORTÁLOK – HTML SCRAPE
# ============================================================

def _strip_html(text):
    """Eltávolítja a HTML tageket és tisztítja a whitespace-t."""
    # HTML entitások visszafordítása
    text = text.replace("&amp;", "&").replace("&nbsp;", " ")
    text = text.replace("&quot;", '"').replace("&#8211;", "–").replace("&#8217;", "'")
    # HTML tagek eltávolítása
    text = re.sub(r"<[^>]+>", "", text)
    # Több whitespace egybe
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _fetch_html(url, timeout=30):
    """HTML letöltése böngésző-szerű header-rel (anti-bot védelem ellen)."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "hu-HU,hu;q=0.9,en;q=0.8",
    }
    resp = requests.get(url, headers=headers, timeout=timeout)
    resp.raise_for_status()
    return resp.text


def fetch_welovebudapest():
    """We Love Budapest gasztro rovat - friss cikkek a kezdőoldalról."""
    try:
        html = _fetch_html("https://welovebudapest.com/gasztro/")
    except Exception as e:
        print(f"  ✗ Hiba a We Love Budapest letöltésénél: {e}")
        return []

    articles = []
    # Két pattern-t használunk:
    # 1. ## [Cím](URL) markdown-szerű
    # 2. <h2><a href="URL">Cím</a></h2> HTML
    patterns = [
        re.compile(
            r'\[([^\]]{15,250})\]\((https://welovebudapest\.com/(?:cikk|toplista)/[^\)]+)\)',
            re.IGNORECASE,
        ),
        re.compile(
            r'<a[^>]+href="(https://welovebudapest\.com/(?:cikk|toplista)/[^"]+)"[^>]*>'
            r'\s*<h[2-4][^>]*>([\s\S]{15,300}?)</h[2-4]>',
            re.IGNORECASE,
        ),
    ]

    seen_urls = set()
    for pattern in patterns:
        for match in pattern.finditer(html):
            # Pattern 1: group(1)=title, group(2)=url
            # Pattern 2: group(1)=url, group(2)=title
            if pattern.pattern.startswith(r"\["):
                title = _strip_html(match.group(1))
                url = match.group(2)
            else:
                url = match.group(1)
                title = _strip_html(match.group(2))

            if url in seen_urls:
                continue
            if not title or len(title) < 10:
                continue
            seen_urls.add(url)
            articles.append({"title": title.strip(), "url": url})

    print(f"  ✓ We Love Budapest: {len(articles)} cikk találva")
    return articles


def fetch_funzine():
    """Funzine.hu gasztro rovat - friss cikkek."""
    try:
        html = _fetch_html("https://funzine.hu/category/gasztro/")
    except Exception as e:
        print(f"  ✗ Hiba a Funzine letöltésénél: {e}")
        return []

    articles = []
    # Funzine cikk URL minta: https://funzine.hu/YYYY/MM/DD/gasztro/SLUG/
    # Két pattern, akárcsak WLB-nél
    patterns = [
        re.compile(
            r'\[([^\]]{15,250})\]\((https://funzine\.hu/\d{4}/\d{2}/\d{2}/gasztro/[^\)]+)\)',
            re.IGNORECASE,
        ),
        re.compile(
            r'<a[^>]+href="(https://funzine\.hu/\d{4}/\d{2}/\d{2}/gasztro/[^"]+)"[^>]*'
            r'title="([^"]{15,300})"',
            re.IGNORECASE,
        ),
    ]

    seen_urls = set()
    for pattern in patterns:
        for match in pattern.finditer(html):
            if pattern.pattern.startswith(r"\["):
                title = _strip_html(match.group(1))
                url = match.group(2)
            else:
                url = match.group(1)
                title = _strip_html(match.group(2))

            if url in seen_urls:
                continue
            if not title or len(title) < 10:
                continue
            seen_urls.add(url)
            articles.append({"title": title.strip(), "url": url})

    print(f"  ✓ Funzine: {len(articles)} cikk találva")
    return articles


def fetch_timeout_budapest():
    """Time Out Budapest étterem cikkek."""
    # Több potenciális URL-en is megpróbáljuk
    html = None
    for url in [
        "https://www.timeout.com/hu/budapest/ettermek",
        "https://www.timeout.com/hu/budapest/etterem",
        "https://www.timeout.com/hu/budapest",
    ]:
        try:
            html = _fetch_html(url)
            break
        except Exception:
            continue

    if not html:
        print("  ✗ Hiba a Time Out Budapest letöltésénél: nem érhető el")
        return []

    articles = []
    # Time Out: étterem-kapcsolt URL-eket keresünk
    relevant_keywords = [
        "etterem", "ettermek", "kavezo", "bistro", "bisztro",
        "food", "gasztro", "reggeli", "brunch", "kavehaz",
        "pekseg", "restaurant", "bar"
    ]
    patterns = [
        re.compile(
            r'\[([^\]]{15,250})\]\((https://www\.timeout\.com/hu/budapest/[^\)]+)\)',
            re.IGNORECASE,
        ),
        re.compile(
            r'<a[^>]+href="(https://www\.timeout\.com/hu/budapest/[^"]+)"[^>]*'
            r'(?:title|aria-label)="([^"]{15,300})"',
            re.IGNORECASE,
        ),
    ]

    seen_urls = set()
    for pattern in patterns:
        for match in pattern.finditer(html):
            if pattern.pattern.startswith(r"\["):
                title = _strip_html(match.group(1))
                url = match.group(2)
            else:
                url = match.group(1)
                title = _strip_html(match.group(2))

            # Csak vendéglátós tematikájú cikkek
            if not any(kw in url.lower() for kw in relevant_keywords):
                continue
            if url in seen_urls:
                continue
            if not title or len(title) < 10:
                continue
            seen_urls.add(url)
            articles.append({"title": title.strip(), "url": url})

    print(f"  ✓ Time Out Budapest: {len(articles)} cikk találva")
    return articles


def filter_new_articles(articles, known_urls):
    """Csak az új cikkek, amik még nem voltak ismertek."""
    return [a for a in articles if a.get("url") not in known_urls]


# ============================================================
# ÖNKORMÁNYZATI ÜZLETNYILVÁNTARTÁS – govcenter.hu
# ============================================================

def _parse_govcenter_table(html, keruletnev):
    """
    Kinyeri az üzleteket a govcenter.hu HTML táblázatából.
    A táblázat oszlopai jellemzően: nyilvántartási szám, nyilvántartásba vétel
    dátuma, üzlet neve, cím, üzemeltető, tevékenység.
    Mivel a pontos HTML struktúra változhat, többféle parse-stratégiát próbálunk.
    """
    uzletek = []

    # Fejléc-szavak, amik egyértelműen jelzik hogy ez a fejléc sor (nem valódi üzlet)
    fejlec_szavak = {
        "sorszám", "sorszam", "nyilv.szám", "nyilv. szám", "nyilv.szam", "nyilvszam",
        "üzlet", "uzlet", "kereskedő", "kereskedo", "tevékenység", "tevekenyseg",
        "név", "nev", "cím", "cim", "üzemeltető", "uzemelteto",
        "dátum", "datum", "típus", "tipus",
    }

    # Stratégia: minden <tr> sort kiveszünk, és a cellákat <td>-nként
    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.IGNORECASE | re.DOTALL)
    for row in rows:
        cells = re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", row, re.IGNORECASE | re.DOTALL)
        if len(cells) < 3:
            continue
        # Tisztítjuk a cellákat
        clean = [_strip_html(c) for c in cells]
        # Kihagyjuk az üres sorokat
        if not any(clean):
            continue

        # Heurisztika: ez egy fejléc sor?
        # Ha minden cella rövid (max 25 karakter) és van benne fejléc-szó → fejléc
        ossz_szoveg = " ".join(clean).lower()
        is_fejlec = (
            all(len(c) <= 25 for c in clean)
            and any(fsz in ossz_szoveg for fsz in fejlec_szavak)
            and not re.search(r"\d{4}", ossz_szoveg)  # ha NINCS benne év, akkor inkább fejléc
        )
        if is_fejlec:
            continue

        # Heurisztika: keresünk dátumot valamelyik cellában (YYYY.MM.DD vagy YYYY-MM-DD)
        datum = ""
        for c in clean:
            m = re.search(r"(20\d{2})[.\-/]\s?(\d{1,2})[.\-/]\s?(\d{1,2})", c)
            if m:
                datum = f"{m.group(1)}.{m.group(2).zfill(2)}.{m.group(3).zfill(2)}"
                break

        # Az üzlet neve és címe jellemzően a leghosszabb szöveges cellák
        szoveges = [c for c in clean if len(c) > 2 and not re.fullmatch(r"[\d.\-/\s]+", c)]
        if not szoveges:
            continue

        # Még egy szűrés: ha az összes szöveges cella csak fejléc-szó, kihagyjuk
        if all(c.lower() in fejlec_szavak for c in szoveges):
            continue

        uzlet = {
            "kerulet": keruletnev,
            "datum": datum,
            "cellak": clean,
            "nev_cim": " | ".join(szoveges[:4]),
        }
        # Egyedi kulcs: a sor teljes tartalma
        uzlet["kulcs"] = f"{keruletnev}::{'|'.join(clean)}"
        uzletek.append(uzlet)

    return uzletek


def fetch_govcenter():
    """
    Lekérdezi a govcenter.hu üzletnyilvántartást a beállított kerületekre.
    Mindkét nyilvántartás-típust lekéri:
      tip=1: bejelentés-köteles kereskedelmi tevékenység
      tip=2: működési engedélyhez kötött üzletek
    Így nem marad ki étterem egyik listáról sem.
    """
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "hu-HU,hu;q=0.9",
    }

    TIP_NEVEK = {1: "bejelentés", 2: "működési eng."}
    all_uzletek = []

    for ker in GOVCENTER_KERULETEK:
        ker_total = 0
        for tip in [1, 2]:
            url = (
                f"https://www.govcenter.hu/uzlet/Public/Uzleteklista.aspx"
                f"?tip={tip}&onk_id={ker['onk_id']}"
            )
            try:
                resp = requests.get(url, headers=headers, timeout=30)
                if resp.status_code != 200:
                    print(f"  ✗ {ker['name']} (tip={tip}): HTTP {resp.status_code}")
                    continue
                uzletek = _parse_govcenter_table(resp.text, ker["name"])
                ker_total += len(uzletek)
                all_uzletek.extend(uzletek)
            except Exception as e:
                print(f"  ✗ Hiba {ker['name']} (tip={tip}): {e}")

        if ker_total == 0:
            print(f"  ⚠️  {ker['name']}: 0 sor (tip=1 és tip=2 is üres)")
        else:
            print(f"  ✓ {ker['name']}: {ker_total} üzlet sor (tip=1+2)")

    return all_uzletek

    return all_uzletek


def _is_foodora_relevant(text):
    """
    Eldönti egy üzletsor szövegéről, hogy releváns-e Foodora szempontból.
    Vendéglátós + élelmiszerbolt + pékség, cukrászda.

    Pozitív kulcsszó kell → de negatív nem lehet (ezek tipikus tévedések).

    Fontos: a negatív szavak szó-határosan keresünk (\b szóhatár), hogy ne
    szűrjük ki pl. a "kasmírvirág" cégnevet a "virág" negatív szó miatt.
    """
    t = text.lower()

    # NEGATÍV (kizáró) kulcsszavak — ha bármelyik megvan, NEM releváns
    # SZÓHATÁROS keresés (\b szóhatár) — fontos!
    negativ_szohataros = [
        # Üzemanyag, autó
        r"\btöltőállomás", r"\btoltoallomas", r"\btöltöállomás",
        r"\büzemanyag", r"\bautóalkatrész", r"\balkatrész",
        r"\bautószerel", r"\bgumiszervíz", r"\bautómosó",
        # Barkács, festék, vegyiáru, építőanyag
        r"\bbarkács", r"\bbarkacs", r"\bfesték", r"\bfestek",
        r"\bvegyiáru", r"\bépítőanyag", r"\bepitoanyag",
        r"\bcsempe", r"\bburkolat",
        # Papír-írószer, könyv
        r"\bpapír-írószer", r"\bírószer", r"\biroszer",
        r"\bkönyvesbolt", r"\bkönyvtár",
        # Virág — szóhatáros, hogy a "Kasmírvirág" ne legyen kizárva!
        r"\bvirág\b", r"\bvirágüzlet", r"\bvirágbolt",
        # Ruha, divat, cipő (szóhatáros)
        r"\bruházat", r"\bruha\b", r"\bdivat", r"\bcipő\b",
        r"\böltöny", r"\bhasználtruha",
        # Műszaki, elektronika
        r"\belektronika", r"\bműszaki", r"\bmuszaki",
        r"\bszámítástechnika", r"\binformatika",
        # Csomagküldő, posta
        r"\bcsomagküldő", r"\bcsomagkuldo",
        # Mozgóbolt
        r"\bmozgóbolt", r"\bmozgobolt",
        # Egyéb szolgáltatás
        r"\bfodrász", r"\bfodrasz", r"\bkozmetika",
        r"\bmasszázs", r"\bmasszazs", r"\bszolárium", r"\bmanikűr",
        # Játék, sport, optika
        r"\bjátékterem", r"\bjátékbolt", r"\bjáték-", r"\bjatek-",
        r"\bsportbolt", r"\bsportszer", r"\boptika",
        # Dohány, drogéria
        r"\btrafik", r"\bdohánybolt", r"\bdohany", r"\bdrogéria",
        # Áruház, vegyesbolt (de NE szűrjön ha élelmiszerre is utal)
        r"\btedi\b", r"\blidl áruház", r"\bauchan\b", r"\bikea\b",
        # Tárgyak
        r"\bbútor", r"\bbutor",
        # Poszter, dekoráció
        r"\bposzter", r"\bdekoráció",
    ]
    for pattern in negativ_szohataros:
        if re.search(pattern, t):
            return False

    # POZITÍV kulcsszavak — legalább 1 kell
    # Itt nem kell szóhatár, mert ezek a szavak ritkán részei másiknak
    pozitiv = [
        # Vendéglátás — éttermek
        "étterem", "etterem", "étkezde", "etkezde", "kifőzde", "kifozde",
        "vendéglő", "vendeglo", "bisztró", "bisztro", "bistro",
        "étkező", "etkezo", "falatozó", "falatozo",
        "büfé", "bufe", "büffe", "büffé", "gyorsétterem",
        "kantin", "csárda", "csarda",
        # Kocsma, bár, sör
        "kocsma", "söröző", "sörözõ", "sorozo", "pub",
        " sör", " sor", "beer", "beerhouse", "ale",
        "koktélbár", "koktelbar", "koktailbar", "koktail", "koktél",
        "cocktail", "coctail",  # FéLIX Coctails miatt
        "borozó", "borozo", "borbar", "borbár",
        # Kávé, cukrászda
        "kávézó", "kavezo", "kávéház", "kavehaz", "kávéháza",
        "cukrászda", "cukraszda", "konditorei",
        "fagylaltozó", "fagyizó", "fagylalt",
        "pékség", "pekseg", "pék ", "bakery",
        # Konyhatípusok
        "pizza", "pizzéria", "pizzeria",
        "kebab", "gyros", "gyrosos", "burger", "hamburger",
        "ramen", "sushi", "thai", "kínai", "vietnami",
        "taco", "burrito", "mexikói",
        "falafel", "sandwich", "szendvics",
        "hot dog", "hotdog",
        # Élelmiszer
        "élelmiszer", "elelmiszer", "élelmiszerbolt", "elelmiszerbolt",
        "élelm.", "élelmiszerüzlet",
        "kisbolt", "kis-bolt", "minibolt",  # Zacc platz miatt
        "csemege", "delikát",
        "biobolt", "bio bolt", "bio-bolt", "bio-világ", "biopiac",
        "biouzlet", "bio üzlet",
        "zöldséges", "zoldseges", "zöldség-gyümölcs",
        "halbolt", "hal-bolt", "halpiac",
        "mészárszék", "meszarszek", "húsbolt", "húsüzlet", "hentes",
        "fűszerbolt", "fuszerbolt",
        # Kávé alapjából (utótaggal, hogy ne legyen téves egyezés)
        "kávé ", " kávé", "kave ",
        # Hostel/kávéház/teaház
        "teaház", "teahaz",
        # Egyéb vendéglátós formák
        "gasztro", " food", "restaurant", "ristorante",
        "trattoria", "osteria", "pizzaholic",
    ]
    for p in pozitiv:
        if p in t:
            return True
    return False


def filter_new_govcenter(uzletek, known_govcenter_dict):
    """Csak az új ÉS Foodora-releváns ÉS friss üzletsorok."""
    DATUM_CUTOFF = "2026.01.01"
    result = []
    for u in uzletek:
        if u.get("kulcs") in known_govcenter_dict:
            continue
        datum = u.get("datum", "")
        if datum and datum < DATUM_CUTOFF:
            continue
        text = u.get("nev_cim", "") + " " + " ".join(u.get("cellak", []))
        if not _is_foodora_relevant(text):
            continue
        result.append(u)
    return result


# ============================================================
# E-MAIL FORMÁZÁS ÉS KÜLDÉS
# ============================================================

def format_email(new_places, new_domains, news,
                 new_wlb_articles, new_funzine_articles, new_timeout_articles,
                 new_govcenter=None):
    """HTML e-mail összeállítása az új találatokkal."""
    today = datetime.now().strftime("%Y-%m-%d")
    if new_govcenter is None:
        new_govcenter = []

    html = f"""
    <html>
    <head>
    <style>
      body {{ font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto; color: #333; }}
      h1 {{ color: #d70f64; border-bottom: 3px solid #d70f64; padding-bottom: 10px; }}
      h2 {{ color: #d70f64; margin-top: 30px; }}
      .place {{ background: #f9f9f9; padding: 12px; margin: 8px 0; border-left: 4px solid #d70f64; border-radius: 4px; }}
      .place.future {{ border-left-color: #ff9500; background: #fff8e6; }}
      .name {{ font-weight: bold; font-size: 16px; }}
      .badge {{ display: inline-block; padding: 2px 8px; border-radius: 3px; font-size: 11px; font-weight: bold; }}
      .badge.future {{ background: #ff9500; color: white; }}
      .badge.new {{ background: #d70f64; color: white; }}
      .meta {{ color: #666; font-size: 13px; margin-top: 4px; }}
      a {{ color: #d70f64; text-decoration: none; }}
      .stats {{ background: #f0f0f0; padding: 10px; border-radius: 5px; }}
      .empty {{ color: #999; font-style: italic; }}
    </style>
    </head>
    <body>
    <h1>🍕 Foodora Lead Finder – {today}</h1>
    <div class="stats">
      <strong>Mai találatok:</strong>
      Google Maps: {len(new_places)} új hely &nbsp;|&nbsp;
      Új domainek: {len(new_domains)} &nbsp;|&nbsp;
      Hírek: {len(news)} &nbsp;|&nbsp;
      WLB: {len(new_wlb_articles)} &nbsp;|&nbsp;
      Funzine: {len(new_funzine_articles)} &nbsp;|&nbsp;
      Time Out: {len(new_timeout_articles)} &nbsp;|&nbsp;
      🏛️ Önkormányzat: {len(new_govcenter)}
    </div>
    """

    # KIEMELT: önkormányzati üzletbejelentések (a legértékesebb forrás, ezért legfelül)
    html += "<h2>🏛️ Új önkormányzati üzletbejelentések (KIEMELT)</h2>"
    if not new_govcenter:
        html += '<p class="empty">Ma nem volt új önkormányzati bejelentés (vagy a forrás épp nem elérhető).</p>'
    else:
        html += ('<p style="color:#666;font-size:13px;">'
                 'Frissen bejelentett kereskedelmi/vendéglátó egységek a kerületi '
                 'nyilvántartásokból. Ezek gyakran még nincsenek fent a Google Maps-en!</p>')
        for u in new_govcenter[:60]:
            datum = u.get("datum", "")
            html += f"""
            <div class="place future">
              <div class="name">{u.get('nev_cim', '')}</div>
              <div class="meta">📍 {u.get('kerulet', '')}{' • 📅 ' + datum if datum else ''}</div>
            </div>
            """
        if len(new_govcenter) > 60:
            html += f"<p><em>... és még {len(new_govcenter)-60} bejelentés</em></p>"

    # 1. ÚJ HELYEK GOOGLE MAPS-RŐL
    html += "<h2>📍 Új helyek a Google Maps-en</h2>"
    if not new_places:
        html += '<p class="empty">Ma nem találtunk új helyet.</p>'
    else:
        # Először a future opening, mert ezek a legértékesebbek
        future = [p for p in new_places if p.get("businessStatus") == "FUTURE_OPENING"]
        current = [p for p in new_places if p.get("businessStatus") != "FUTURE_OPENING"]

        if future:
            html += "<h3>🔥 Hamarosan nyíló helyek (TOP PRIORITÁS!)</h3>"
            for p in future:
                html += format_place_html(p, is_future=True)

        if current:
            html += "<h3>✨ Frissen megjelent helyek</h3>"
            for p in current[:50]:  # max 50, nehogy szétessen az e-mail
                html += format_place_html(p, is_future=False)

    # 2. ÚJ DOMAINEK
    html += "<h2>🌐 Új weboldalak (vendéglátós kulcsszavakkal)</h2>"
    if not new_domains:
        html += '<p class="empty">Ma nem találtunk új domaint.</p>'
    else:
        html += "<ul>"
        for d in sorted(new_domains)[:50]:
            html += f'<li><a href="https://{d}">{d}</a></li>'
        html += "</ul>"
        if len(new_domains) > 50:
            html += f"<p><em>... és még {len(new_domains)-50} domain</em></p>"

    # 3. HÍREK
    html += "<h2>📰 Friss hírek</h2>"
    if not news:
        html += '<p class="empty">Ma nem volt releváns hír.</p>'
    else:
        for n in news[:20]:
            html += f"""
            <div class="place">
              <div class="name"><a href="{n['link']}">{n['title']}</a></div>
              <div class="meta">{n.get('source', '')} • {n.get('published', '')}</div>
            </div>
            """

    # 4. WE LOVE BUDAPEST CIKKEK
    html += "<h2>🌶️ We Love Budapest – új cikkek a gasztro rovatból</h2>"
    if not new_wlb_articles:
        html += '<p class="empty">Ma nem volt új cikk.</p>'
    else:
        for a in new_wlb_articles[:25]:
            html += f"""
            <div class="place">
              <div class="name"><a href="{a['url']}">{a['title']}</a></div>
              <div class="meta">welovebudapest.com</div>
            </div>
            """

    # 5. FUNZINE CIKKEK
    html += "<h2>🎉 Funzine.hu – új cikkek a gasztro rovatból</h2>"
    if not new_funzine_articles:
        html += '<p class="empty">Ma nem volt új cikk.</p>'
    else:
        for a in new_funzine_articles[:25]:
            html += f"""
            <div class="place">
              <div class="name"><a href="{a['url']}">{a['title']}</a></div>
              <div class="meta">funzine.hu</div>
            </div>
            """

    # 6. TIME OUT BUDAPEST CIKKEK
    html += "<h2>🌍 Time Out Budapest – új éttermes cikkek</h2>"
    if not new_timeout_articles:
        html += '<p class="empty">Ma nem volt új cikk.</p>'
    else:
        for a in new_timeout_articles[:25]:
            html += f"""
            <div class="place">
              <div class="name"><a href="{a['url']}">{a['title']}</a></div>
              <div class="meta">timeout.com</div>
            </div>
            """

    html += """
    <hr>
    <p style="color: #999; font-size: 11px;">
      Ezt az e-mailt a Foodora Lead Finder automata küldte.
      Adatforrások: Google Places API, crt.sh, Google News, We Love Budapest, Funzine, Time Out Budapest, önkormányzati üzletnyilvántartás (govcenter.hu).
    </p>
    </body>
    </html>
    """
    return html


def format_place_html(p, is_future=False):
    name = p.get("displayName", {}).get("text", "Névtelen")
    address = p.get("formattedAddress", "")
    maps_url = p.get("googleMapsUri", "")
    website = p.get("websiteUri", "")
    phone = p.get("nationalPhoneNumber", "")
    place_type = p.get("primaryType", "")
    area = p.get("_search_area", "")

    badge = ""
    extra = ""
    if is_future:
        badge = '<span class="badge future">HAMAROSAN NYÍL</span>'
        opening = p.get("openingDate", {})
        if opening:
            y = opening.get("year", "")
            m = opening.get("month", "")
            d = opening.get("day", "")
            try:
                if y and m and d:
                    extra = f"<div class='meta'>📅 Várható nyitás: {y}-{int(m):02d}-{int(d):02d}</div>"
                else:
                    extra = ""
            except (ValueError, TypeError):
                extra = ""           
    else:
        badge = '<span class="badge new">ÚJ</span>'

    css_class = "place future" if is_future else "place"

    html = f'<div class="{css_class}">'
    html += f'<div class="name">{badge} {name}</div>'
    if address:
        html += f'<div class="meta">📍 {address}</div>'
    if place_type:
        html += f'<div class="meta">🏷️ {place_type} • {area}</div>'
    if phone:
        html += f'<div class="meta">📞 {phone}</div>'
    if website:
        html += f'<div class="meta">🌐 <a href="{website}">{website}</a></div>'
    if extra:
        html += extra
    if maps_url:
        html += f'<div class="meta"><a href="{maps_url}">▶ Megnyit Google Maps-ben</a></div>'
    html += "</div>"
    return html


def send_email(html_content):
    """E-mail küldés Gmail SMTP-n keresztül."""
    if not (EMAIL_FROM and EMAIL_TO and EMAIL_PASSWORD):
        print("⚠️  E-mail credentials hiányoznak, kihagyom a küldést.")
        print("=" * 60)
        print("HTML ELŐNÉZET (első 2000 karakter):")
        print(html_content[:2000])
        return

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"🍕 Foodora Lead Finder – {datetime.now().strftime('%Y-%m-%d')}"
    msg["From"] = EMAIL_FROM
    msg["To"] = EMAIL_TO
    msg.attach(MIMEText(html_content, "html", "utf-8"))

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, context=context) as server:
        server.login(EMAIL_FROM, EMAIL_PASSWORD)
        server.send_message(msg)
    print(f"✉️  E-mail elküldve: {EMAIL_TO}")


def generate_full_database_xlsx(known_places, known_govcenter):
    """
    Exportálja a TELJES ismert helyek adatbázisát xlsx-be.
    - Google Places: minden ismert hely (névvel, címmel, telefonnal ahol van)
    - Govcenter: csak a Foodora-releváns bejegyzések
    Napról napra egyre teljesebb lesz ahogy a helyek újra előkerülnek.
    """
    wb = Workbook()

    # ── 1. lap: Google Places ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Google Places"

    header_fill = PatternFill("solid", start_color="1A5276")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    data_font   = Font(name="Arial", size=9)

    headers1 = ["Felfedezve", "Név", "Cím", "Telefon", "Típus", "Google Maps link"]
    widths1   = [15, 35, 45, 18, 20, 55]
    for col, (h, w) in enumerate(zip(headers1, widths1), 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.column_dimensions[cell.column_letter].width = w
    ws1.row_dimensions[1].height = 20
    ws1.freeze_panes = "A2"

    def sort_key(item):
        details = item[1]
        if details and details.get("name"):
            # Teljes adatú sorok elöl, legújabb felfedezés elöl
            date = details.get("discovered_at") or "0000-00-00"
            return (0, date)
        return (1, "")

    # Rendezés: teljes adatú sorok elöl (legújabb elöl), hiányos sorok hátul
    sorted_places = sorted(known_places.items(), key=sort_key, reverse=False)
    # A dátum szerint fordítva akarjuk (legújabb elöl) de csak az első csoportban
    with_data = [(k,v) for k,v in sorted_places if v and v.get("name")]
    without_data = [(k,v) for k,v in sorted_places if not (v and v.get("name"))]
    with_data.sort(key=lambda x: x[1].get("discovered_at") or "0000", reverse=True)
    sorted_places = with_data + without_data
    row = 2
    fill_even = PatternFill("solid", start_color="D6EAF8")
    fill_odd  = PatternFill("solid", start_color="FFFFFF")
    fill_empty = PatternFill("solid", start_color="F2F3F4")

    for pid, details in sorted_places:
        has_data = details and details.get("name")
        fill = fill_empty if not has_data else (fill_even if row % 2 == 0 else fill_odd)
        if has_data:
            vals = [
                details.get("discovered_at") or "?",
                details.get("name", ""),
                details.get("address", ""),
                details.get("phone", ""),
                details.get("type", ""),
                f"https://maps.google.com/?cid={pid}" if pid else "",
            ]
        else:
            vals = ["ismeretlen", "", "", "", "", ""]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = data_font
            cell.alignment = Alignment(vertical="center")
        row += 1

    ws1.auto_filter.ref = f"A1:F{row-1}"

    # ── 2. lap: Önkormányzati bejelentések ───────────────────
    ws2 = wb.create_sheet("Önkormányzat")
    headers2 = ["Felfedezve", "Bejelentés dátuma", "Kerület", "Üzlet neve / Cím"]
    widths2   = [15, 20, 28, 80]
    for col, (h, w) in enumerate(zip(headers2, widths2), 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[cell.column_letter].width = w
    ws2.row_dimensions[1].height = 20
    ws2.freeze_panes = "A2"

    row2 = 2
    fill_gov_even = PatternFill("solid", start_color="D5F5E3")
    fill_gov_odd  = PatternFill("solid", start_color="FFFFFF")

    # Csak Foodora-releváns govcenter bejegyzések, dátum szerint rendezve
    gov_entries = []
    for kulcs, details in known_govcenter.items():
        if details is None:
            continue
        nev_cim = details.get("nev_cim", "")
        if not _is_foodora_relevant(nev_cim):
            continue
        gov_entries.append(details)

    gov_entries.sort(key=lambda d: d.get("datum") or "0000", reverse=True)

    for details in gov_entries:
        fill = fill_gov_even if row2 % 2 == 0 else fill_gov_odd
        vals = [
            details.get("discovered_at") or "?",
            details.get("datum", ""),
            details.get("kerulet", ""),
            details.get("nev_cim", ""),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row2, column=col, value=val)
            cell.fill = fill
            cell.font = data_font
            cell.alignment = Alignment(vertical="center")
        row2 += 1

    ws2.auto_filter.ref = f"A1:D{row2-1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), len(known_places), len(gov_entries)


def send_xlsx_email(xlsx_bytes, places_count, gov_count):
    """Elküldi a teljes adatbázis xlsx-et mellékletként külön e-mailben."""
    if not (EMAIL_FROM and EMAIL_TO and EMAIL_PASSWORD):
        print("⚠️  E-mail credentials hiányoznak, xlsx küldés kihagyva.")
        return

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"Foodora_adatbazis_{today}.xlsx"
    total = places_count + gov_count

    msg = MIMEMultipart()
    msg["Subject"] = f"📊 Foodora adatbázis ({today}) – {total} bejegyzés"
    msg["From"] = EMAIL_FROM
    msg["To"] = EMAIL_TO

    body = MIMEText(
        f"Napi adatbázis exportálás – {today}\n\n"
        f"Google Places lap: {places_count} hely\n"
        f"Önkormányzati lap: {gov_count} Foodora-releváns bejegyzés\n\n"
        f"A táblázat napról napra bővül és gazdagodik.\n"
        f"Kék sorok = Google Places | Zöld sorok = Önkormányzat\n"
        f"Szürke sorok = még hiányos adat (idővel feltöltődik)",
        "plain", "utf-8"
    )
    msg.attach(body)

    attachment = MIMEApplication(xlsx_bytes, Name=filename)
    attachment["Content-Disposition"] = f'attachment; filename="{filename}"'
    msg.attach(attachment)

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, context=context) as server:
        server.login(EMAIL_FROM, EMAIL_PASSWORD)
        server.send_message(msg)
    print(f"📊 Adatbázis xlsx elküldve: {EMAIL_TO} ({filename}, {total} sor)")


# ============================================================
# ADATTÁROLÁS - mit láttunk már
# ============================================================

def load_known(path):
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_known(path, items):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(sorted(list(items)), f, ensure_ascii=False, indent=2)


def load_known_places(path):
    """Visszafelé kompatibilis betöltés: lista → dict migráció."""
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return {pid: None for pid in data}
    return data


def save_known_places(path, places_dict):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(places_dict, f, ensure_ascii=False, indent=2)


def load_known_govcenter(path):
    """Visszafelé kompatibilis betöltés: lista → dict migráció."""
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return {kulcs: None for kulcs in data}
    return data


def save_known_govcenter(path, govcenter_dict):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(govcenter_dict, f, ensure_ascii=False, indent=2)


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 60)
    print(f"🚀 Foodora Lead Finder – {datetime.now().isoformat()}")
    print("=" * 60)

    # 1. Google Places
    print("\n📍 Google Places API lekérdezés...")
    known_places = load_known_places(KNOWN_PLACES_FILE)
    all_places = fetch_google_places()
    new_places = filter_new_places(all_places, set(known_places.keys()))
    print(f"  → {len(new_places)} ÚJ hely (összesen ismert: {len(known_places)})")

    # 2. Új domainek
    print("\n🌐 crt.sh új domainek lekérdezése...")
    known_domains = load_known(KNOWN_DOMAINS_FILE)
    found_domains = fetch_new_domains()
    new_domains = [d for d in found_domains if d not in known_domains]
    print(f"  → {len(new_domains)} ÚJ domain")

    # 3. Hírek
    print("\n📰 Google News RSS...")
    news = fetch_news()
    print(f"  → {len(news)} friss hír")

    # 4. We Love Budapest gasztro cikkek
    print("\n🌶️  We Love Budapest gasztro cikkek...")
    known_wlb = load_known(KNOWN_WLB_FILE)
    wlb_articles = fetch_welovebudapest()
    new_wlb_articles = filter_new_articles(wlb_articles, known_wlb)
    print(f"  → {len(new_wlb_articles)} ÚJ cikk (összesen ismert: {len(known_wlb)})")

    # 5. Funzine gasztro cikkek
    print("\n🎉 Funzine gasztro cikkek...")
    known_funzine = load_known(KNOWN_FUNZINE_FILE)
    funzine_articles = fetch_funzine()
    new_funzine_articles = filter_new_articles(funzine_articles, known_funzine)
    print(f"  → {len(new_funzine_articles)} ÚJ cikk (összesen ismert: {len(known_funzine)})")

    # 6. Time Out Budapest cikkek
    print("\n🌍 Time Out Budapest cikkek...")
    known_timeout = load_known(KNOWN_TIMEOUT_FILE)
    timeout_articles = fetch_timeout_budapest()
    new_timeout_articles = filter_new_articles(timeout_articles, known_timeout)
    print(f"  → {len(new_timeout_articles)} ÚJ cikk (összesen ismert: {len(known_timeout)})")

    # 7. Önkormányzati üzletnyilvántartás (govcenter.hu)
    print("\n🏛️  Önkormányzati üzletnyilvántartás (govcenter.hu)...")
    known_govcenter = load_known_govcenter(KNOWN_GOVCENTER_FILE)
    govcenter_uzletek = fetch_govcenter()
    new_govcenter = filter_new_govcenter(govcenter_uzletek, known_govcenter)
    print(f"  → {len(new_govcenter)} ÚJ üzletsor (összesen ismert: {len(known_govcenter)})")

    # 8. E-mail (HTML összefoglaló)
    print("\n✉️  E-mail összeállítása és küldése...")
    html = format_email(
        new_places, new_domains, news,
        new_wlb_articles, new_funzine_articles, new_timeout_articles,
        new_govcenter,
    )
    send_email(html)

    # 9. Xlsx melléklet: a mentés UTÁN küldjük, hogy az aznap pótolt adatok is benne legyenek

    # 10. Mentés + hiányzó adatok pótlása
    today_str = datetime.now().strftime("%Y-%m-%d")
    enriched = 0
    for p in all_places:
        pid = p.get("id")
        if not pid:
            continue
        name = p.get("displayName", {}).get("text", "")
        address = p.get("formattedAddress", "")
        phone = p.get("nationalPhoneNumber", "")
        ptype = p.get("primaryType", "")

        if pid not in known_places:
            # Új hely: teljes adattal mentjük
            known_places[pid] = {
                "name": name,
                "address": address,
                "phone": phone,
                "type": ptype,
                "discovered_at": today_str,
            }
        else:
            # Már ismert hely: ha hiányoznak az adatok, pótoljuk
            existing = known_places[pid]
            if existing is None or not (existing or {}).get("name"):
                # Régi lista-formátum migráció VAGY hiányzó adatok
                original_date = (existing or {}).get("discovered_at") if existing else None
                known_places[pid] = {
                    "name": name,
                    "address": address,
                    "phone": phone,
                    "type": ptype,
                    "discovered_at": original_date,  # eredeti felfedezési dátum megtartása
                }
                enriched += 1
    if enriched:
        print(f"  ✓ {enriched} régi bejegyzés adatai pótolva")
    # Govcenter részletek mentése
    for u in govcenter_uzletek:
        kulcs = u.get("kulcs")
        if not kulcs:
            continue
        if kulcs not in known_govcenter or known_govcenter[kulcs] is None:
            known_govcenter[kulcs] = {
                "nev_cim": u.get("nev_cim", ""),
                "datum": u.get("datum", ""),
                "kerulet": u.get("kerulet", ""),
                "discovered_at": today_str,
            }
    known_domains.update(found_domains)
    known_wlb.update(a["url"] for a in wlb_articles)
    known_funzine.update(a["url"] for a in funzine_articles)
    known_timeout.update(a["url"] for a in timeout_articles)
    save_known_places(KNOWN_PLACES_FILE, known_places)
    save_known(KNOWN_DOMAINS_FILE, known_domains)
    save_known(KNOWN_WLB_FILE, known_wlb)
    save_known(KNOWN_FUNZINE_FILE, known_funzine)
    save_known(KNOWN_TIMEOUT_FILE, known_timeout)
    save_known_govcenter(KNOWN_GOVCENTER_FILE, known_govcenter)
    print(
        f"\n💾 Adatok elmentve: {len(known_places)} hely, "
        f"{len(known_domains)} domain, "
        f"{len(known_wlb)} WLB, {len(known_funzine)} Funzine, "
        f"{len(known_timeout)} Time Out, "
        f"{len(known_govcenter)} önkormányzati bejegyzés"
    )

    # 11. Teljes adatbázis xlsx küldése (mentés UTÁN, hogy az aznap pótolt adatok is benne legyenek)
    print("\n📊 Teljes adatbázis xlsx összeállítása és küldése...")
    xlsx_bytes, places_cnt, gov_cnt = generate_full_database_xlsx(known_places, known_govcenter)
    send_xlsx_email(xlsx_bytes, places_cnt, gov_cnt)

    print("✅ Kész!")


if __name__ == "__main__":
    main()
