"""
Foodora Lead Finder — Ismert helyek exportálása
================================================

Exportálja a data/known_places.json tartalmát xlsx fájlba.
Oszlopok: Név, Cím, Telefon, Típus, Felfedezés dátuma

Használat:
  python export_places.py                    # Csak a meglévő adatokat exportálja
  python export_places.py --enrich           # Google API-val dúsítja a hiányzó adatokat
  python export_places.py --enrich --limit 100   # Csak az első 100 hiányzó adatot dúsítja

Figyelem --enrich módban: minden hiányzó bejegyzésnél 1 Google API hívás!
"""

import os
import sys
import json
import time
import argparse
from pathlib import Path
from datetime import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

DATA_DIR = Path("data")
KNOWN_PLACES_FILE = DATA_DIR / "known_places.json"
OUTPUT_FILE = Path("data/ismert_helyek.xlsx")

GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")


def load_known_places():
    """Betölti a known_places.json-t, kezel régi (lista) és új (dict) formátumot is."""
    if not KNOWN_PLACES_FILE.exists():
        print("⚠️  data/known_places.json nem található.")
        return {}
    with open(KNOWN_PLACES_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        print(f"  Régi formátum: {len(data)} place_id (részletek nélkül)")
        return {pid: None for pid in data}
    print(f"  Új formátum: {len(data)} bejegyzés")
    return data


def enrich_place(place_id, api_key):
    """Google Places API-ról lekéri egy hely részleteit a place_id alapján."""
    url = f"https://places.googleapis.com/v1/places/{place_id}"
    headers = {
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": (
            "displayName,formattedAddress,nationalPhoneNumber,primaryType"
        ),
    }
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        if resp.status_code == 200:
            d = resp.json()
            return {
                "name": d.get("displayName", {}).get("text", ""),
                "address": d.get("formattedAddress", ""),
                "phone": d.get("nationalPhoneNumber", ""),
                "type": d.get("primaryType", ""),
                "discovered_at": None,
            }
        return None
    except Exception:
        return None


def build_xlsx(places_dict, output_path):
    """Xlsx fájl létrehozása az ismert helyekből."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ismert helyek"

    # Fejléc formázás
    header_fill = PatternFill("solid", start_color="1F6AA5")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Név", "Cím", "Telefon", "Típus", "Felfedezés dátuma", "Place ID"]
    col_widths = [35, 45, 20, 25, 22, 45]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    # Adatok
    row_fill_even = PatternFill("solid", start_color="EBF5FB")
    data_font = Font(name="Arial", size=10)

    sorted_places = sorted(
        places_dict.items(),
        key=lambda x: (x[1] or {}).get("discovered_at") or "0000",
        reverse=True,  # Legfrissebb elöl
    )

    for row_idx, (place_id, details) in enumerate(sorted_places, 2):
        fill = row_fill_even if row_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        if details:
            row_data = [
                details.get("name", ""),
                details.get("address", ""),
                details.get("phone", ""),
                details.get("type", ""),
                details.get("discovered_at") or "ismeretlen",
                place_id,
            ]
        else:
            row_data = ["", "", "", "", "ismeretlen", place_id]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.font = data_font
            cell.alignment = Alignment(vertical="center")

    # Szűrő és sormagasság
    ws.auto_filter.ref = f"A1:F{len(sorted_places)+1}"
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 16

    wb.save(output_path)
    print(f"✅ Elmentve: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Ismert helyek exportálása xlsx-be")
    parser.add_argument("--enrich", action="store_true",
                        help="Google API-val dúsítja a hiányzó adatokat")
    parser.add_argument("--limit", type=int, default=None,
                        help="Maximálisan dúsítandó bejegyzések száma")
    args = parser.parse_args()

    print("=" * 60)
    print("📊 Ismert helyek exportálása")
    print("=" * 60)

    places = load_known_places()
    if not places:
        print("Nincs adat, kilépés.")
        return

    total = len(places)
    has_details = sum(1 for v in places.values() if v and v.get("name"))
    missing = total - has_details
    print(f"  Összesen: {total} hely")
    print(f"  Részletekkel: {has_details}")
    print(f"  Részletek nélkül (régi): {missing}")

    if args.enrich and missing > 0:
        if not GOOGLE_API_KEY:
            print("❌ GOOGLE_API_KEY hiányzik! Dúsítás kihagyva.")
        else:
            limit = args.limit or missing
            to_enrich = [(pid, v) for pid, v in places.items()
                         if not (v and v.get("name"))][:limit]
            print(f"\n🔍 Dúsítás {len(to_enrich)} helyre (Google API)...")
            enriched = 0
            for i, (pid, _) in enumerate(to_enrich, 1):
                details = enrich_place(pid, GOOGLE_API_KEY)
                if details:
                    places[pid] = details
                    enriched += 1
                if i % 10 == 0:
                    print(f"  {i}/{len(to_enrich)} feldolgozva...")
                time.sleep(0.2)  # 5 req/sec, kíméletes tempó
            print(f"  ✓ {enriched} hely dúsítva")

    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    build_xlsx(places, OUTPUT_FILE)

    with_details = sum(1 for v in places.values() if v and v.get("name"))
    print(f"\n📋 {total} hely exportálva, ebből {with_details} teljes adattal")
    print(f"   Fájl: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
